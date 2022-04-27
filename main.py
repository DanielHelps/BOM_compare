import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np


def find_parent(levels_list: list, current_index: int) -> int:
    """Finds the parent of a certain index and returns the parent's index

    :param levels_list: list of levels of BOM
    :type levels_list: list
    :param current_index: index of item whose parent you are looking for
    :type current_index: int
    :return: index of parent
    :rtype: int
    """
    current_level = levels_list[current_index]
    iterated_levels = levels_list[0:current_index]
    # Checking for the current level and up
    iterated_levels = iterated_levels[::-1]
    for i, level in enumerate(iterated_levels):
        if level < current_level:
            # Returning index of parent
            return current_index - i-1


def find_direct_children(levels_list: list, current_index: int) -> list:
    """Finds all the direct children of an item (1 layer down) and returns list of indices for them

    :param levels_list: list of levels of BOM
    :type levels_list: list
    :param current_index: index of parent
    :type current_index: int
    :return: list of indices of direct children
    :rtype: list
    """
    children = []
    current_level = levels_list[current_index]
    for i, level in enumerate(levels_list[current_index+1:]):
        # checking until getting to the next item on the same level
        if level <= current_level:
            break
        elif level is (current_level+1):
            children.append(current_index+i+1)
    return children


def find_all_children(levels_list: list, current_index: int) -> list:
    """Finds all the children of an item (all layer down) and returns list of indices for them

    :param levels_list: list of levels of BOM
    :type levels_list: list
    :param current_index: index of parent
    :type current_index: int
    :return: list of indices of direct children
    :rtype: list
    """
    children = []
    current_level = levels_list[current_index]
    for i, level in enumerate(levels_list[current_index + 1:]):
        if level <= current_level:
            break
        else:
            children.append(current_index + i + 1)
    return children


def find_pn_indices(pns_list: list, required_pn: str) -> list:
    """Finds all the indices of the instances of item P/N,
    for example 36-BEAM-0065 will give indices of all places where
    the item appears

    :param pns_list: list of P/N's to search in
    :type pns_list: list
    :param required_pn: P/N you are looking for
    :type required_pn: str
    :return: list of indices of appearances in pns_list
    :rtype: list
    """
    indices = []
    for i, pn in enumerate(pns_list):
        if pn == required_pn:
            indices.append(i)
    return indices


def did_all_children_matched(match_indices: np.array, levels_list: list) -> np.array:
    """A function to check if all children of a parent matched, and if it did then
    add the parent to the matched indices

    :param match_indices: list of matched items indices
    :type match_indices: np.array
    :param levels_list: list of levels of BOM
    :type levels_list: list
    :return: updated match_indices array
    :rtype: np.array
    """
    for i, level in enumerate(levels_list):
        # find children indices for item
        children_indices = find_all_children(levels_list, len(levels_list)-i-1)
        # If item has children
        if len(children_indices) != 0:
            # start by assuming all children matched
            all_children_matched = True
            for child in children_indices:
                # If there is a child that didn't match
                if child not in match_indices:
                    all_children_matched = False
                    break
            if all_children_matched is True:
                match_indices = np.append(match_indices, len(levels_list)-i-1)
    return match_indices


def find_match_indices(bom_list: list, compare_list: list, levels_list: list) -> np.array:
    """Find all the matches between a BOM and the compared list.
    Match can be gained in 3 ways:
    -Item is directly specified in compare_list
    -Item's father is specified in compare_list
    -All children of the item are specified in compare_list

    :param bom_list: list of bom with P/N's
    :type bom_list: list
    :param compare_list: list of P/N's to compare with bom_list
    :type compare_list: list
    :param levels_list: list of levels of BOM
    :type levels_list: list
    :return: matches indices for bom_list
    :rtype: np.array
    """
    match_indices = np.array([], dtype="int16")
    for compare_pn in compare_list:
        # Find where the compare P/N exists in BOM
        pn_indices = find_pn_indices(bom_list, compare_pn)
        # Add match indices for directly specified items
        match_indices = np.append(match_indices, pn_indices) if pn_indices is not None else []
        # Add matches indices for all children
        for pn_index in pn_indices:
            match_indices = np.append(match_indices, find_all_children(levels_list, pn_index)) if \
                find_all_children(levels_list, pn_index) is not None else []
    # Add matches indices for all parent's indices whose all their children were matched
    match_indices = did_all_children_matched(match_indices, levels_list)
    # Remove duplicates
    match_indices = np.unique(match_indices)

    return match_indices


def create_bom_comparison(bom_path: str, match_indices: np.array):
    """Create a file with all the matches

    :param bom_path: directory path to bom_list
    :type bom_path: str
    :param match_indices: all match indices
    :type match_indices: np.array
    :return: Nothing, creates a file with the matches
    :rtype: .xlsx file
    """
    match_indices = match_indices.astype("int")
    wb = load_workbook(os.path.dirname(bom_path)+"\\BOM.xlsx")
    ws = wb.worksheets[0]
    # Insert column to the left of everything
    ws.insert_cols(1)
    ws["A1"] = "Is match?"
    # Add "Match!" if the match index exists
    for index in match_indices:
        ws[f"A{index + 2}"] = "Match!"
    pass
    wb.save(os.path.dirname(bom_path)+'\\Compare_results.xlsx')
    wb.close()


def main():
    bom_path = "D:\\BOM\\BOM.xlsx"
    compare_path = "D:\\BOM\\Compare.xlsx"
    bom = pd.read_excel(bom_path, sheet_name=0)
    compare = pd.read_excel(compare_path, sheet_name=0)
    level_list = bom['רמת מוצר'].tolist()
    level_list = [int(level.replace(".", "")) for level in level_list]
    compare_pns = compare["מק'ט"].tolist()
    pns_list = bom["מק'ט"].tolist()

    # trial_index = 50
    # print(find_father(level_list, trial_index))
    # print(find_direct_children(level_list, trial_index))
    # print(find_all_children(level_list, trial_index))
    # print(find_pn_indices(pns_list, compare_pns[1]))

    match_indices = find_match_indices(pns_list, compare_pns, level_list)
    create_bom_comparison(bom_path, match_indices)
    pass


if __name__ == '__main__':
    main()
